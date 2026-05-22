import openpyxl, unicodedata, re, datetime

def normalize_name(name):
    if not name: return ''
    n = unicodedata.normalize('NFD', str(name))
    n = ''.join(c for c in n if unicodedata.category(c) != 'Mn')
    return n.lower().strip()

def clean_text(text):
    if not isinstance(text, str):
        return ""
    text = text.lower()
    text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[^a-z0-9\s]', ' ', text)
    return ' '.join(text.split())

def names_match(a, b):
    na, nb = normalize_name(a), normalize_name(b)
    if na == nb: return True
    wa = set(w for w in na.split() if len(w) > 2)
    wb_set = set(w for w in nb.split() if len(w) > 2)
    if not wa or not wb_set: return False
    return len(wa & wb_set) >= min(2, min(len(wa), len(wb_set)))

def task_names_match(cron_task, master_task):
    if not cron_task or not master_task: return False
    def normalize_task_name(name):
        cleaned = clean_text(name)
        if "conciliacion de pasarelas" in cleaned:
            return "conciliacion de pasarelas"
        if "revision de billetera" in cleaned or "billetera usuarios" in cleaned:
            return "revision de billetera usuarios pdv"
        if "revision de eventos" in cleaned or "revision de evento" in cleaned:
            return "revision de eventos"
        return cleaned
    
    norm_cron = normalize_task_name(cron_task)
    norm_master = normalize_task_name(master_task)
    return norm_cron == norm_master or norm_master in norm_cron or norm_cron in norm_master

def set_names_match(set1, set2):
    if not set1 or not set2: return False
    return clean_text(set1) == clean_text(set2) or clean_text(set1) in clean_text(set2) or clean_text(set2) in clean_text(set1)

# Excel date helper
def excel_to_datetime(serial):
    if not serial: return None
    try:
        serial = float(serial)
        return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=serial)
    except:
        return None

def is_same_date(excel_date, js_date):
    if not excel_date or not js_date: return False
    return (excel_date.day == js_date.day and 
            excel_date.month == js_date.month and 
            excel_date.year == js_date.year)

def get_shift_for_date(rows_h, all_blocks, gestor_name, target_date):
    target_block = None
    target_col_index = -1
    
    for block in all_blocks:
        start_row = block['startRow']
        date_row = rows_h[start_row]
        for c in range(1, len(date_row)):
            serial = date_row[c]
            cell_date = excel_to_datetime(serial)
            if cell_date and is_same_date(cell_date, target_date):
                target_block = block
                target_col_index = c
                break
        if target_block: break
        
    if not target_block:
        target_block = all_blocks[-1]
        day_names = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
        day_row = rows_h[target_block['startRow'] + 1]
        js_weekday = (target_date.weekday() + 1) % 7
        target_day_name = day_names[js_weekday]
        
        for c in range(1, len(day_row)):
            day_name = str(day_row[c] or '').strip()
            if normalize_name(day_name) == normalize_name(target_day_name):
                target_col_index = c
                break
        if target_col_index == -1:
            js_day = js_weekday
            target_col_index = 7 if js_day == 0 else js_day

    block_start_row = target_block['startRow']
    for r_idx in range(block_start_row + 2, len(rows_h)):
        r = rows_h[r_idx]
        if not r or not r[0] or str(r[0]).strip() == '' or str(r[0]).strip().upper() == 'GESTOR':
            break
        if names_match(r[0], gestor_name):
            return r[target_col_index] or 'Descansa'
            
    return 'Por Asignar'

# Load files
wb_h = openpyxl.load_workbook('Horario/Horario 2026.xlsx')
ws_h = wb_h.active
rows_h = list(ws_h.values)

# Parse blocks
globalScheduleBlocks = []
r_idx = 0
while r_idx < len(rows_h):
    row = rows_h[r_idx]
    if row and row[0] == 'GESTOR':
        end_row = r_idx + 2
        while end_row < len(rows_h):
            r = rows_h[end_row]
            if not r or not r[0] or str(r[0]).strip() == '' or str(r[0]).strip().upper() == 'GESTOR':
                break
            end_row += 1
        globalScheduleBlocks.append({
            'startRow': r_idx,
            'endRow': end_row - 1
        })
        r_idx = end_row
    else:
        r_idx += 1

wb_c = openpyxl.load_workbook('Cronograma de Tareas/Cronograma Mayo.xlsx')
sheet_names = wb_c.sheetnames

def get_cronograma_columns_for_today(target_date, shift_text):
    js_day = (target_date.weekday() + 1) % 7
    if js_day == 0:
        return [[10, 11]]
    elif js_day == 6:
        return [[7, 8]]
    else:
        return [[1, 2], [4, 5]]

def test_date_assignments(gestor_name, target_date):
    resolved_shift = get_shift_for_date(rows_h, globalScheduleBlocks, gestor_name, target_date)
    col_groups = get_cronograma_columns_for_today(target_date, resolved_shift)
    
    # Load sheet name
    months = {
        "ene": 1, "enero": 1, "feb": 2, "febrero": 2, "mar": 3, "marzo": 3, "abr": 4, "abril": 4,
        "may": 5, "mayo": 5, "jun": 6, "junio": 6, "jul": 7, "julio": 7, "ago": 8, "agosto": 8,
        "sep": 9, "set": 9, "septiembre": 9, "oct": 10, "octubre": 10, "nov": 11, "noviembre": 11,
        "dic": 12, "diciembre": 12
    }

    def parse_sheet_range(sheet_name, year=2026):
        clean = sheet_name.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
        m = re.search(r'Semana\s+\d+\s*-\s*(\d+)\s+(\w+)\s+al\s+(\d+)\s+(\w+)', clean, re.IGNORECASE)
        if m:
            start_day = int(m.group(1))
            start_m_str = m.group(2)[:3].lower()
            end_day = int(m.group(3))
            end_m_str = m.group(4)[:3].lower()
            return datetime.datetime(year, months.get(start_m_str, 1), start_day, 0, 0, 0), datetime.datetime(year, months.get(end_m_str, 1), end_day, 23, 59, 59)
        m = re.search(r'Semana\s+\d+\s*-\s*(\d+)\s+al\s+(\d+)\s+(\w+)', clean, re.IGNORECASE)
        if m:
            start_day = int(m.group(1))
            end_day = int(m.group(2))
            m_str = m.group(3)[:3].lower()
            return datetime.datetime(year, months.get(m_str, 1), start_day, 0, 0, 0), datetime.datetime(year, months.get(m_str, 1), end_day, 23, 59, 59)
        return None

    def get_week_sheet(sheet_names, target_date):
        for name in sheet_names:
            r = parse_sheet_range(name, target_date.year)
            if r:
                start, end = r
                if start <= target_date <= end:
                    return name
        return sheet_names[-1]

    sheet_name = get_week_sheet(sheet_names, target_date)
    ws_c = wb_c[sheet_name]
    rows_c = list(ws_c.values)

    assignments = []
    for col_group in col_groups:
        t_col = col_group[0]
        g_col = col_group[1]
        
        current_set = ""
        for r_idx in range(len(rows_c)):
            row = rows_c[r_idx]
            if not row or len(row) <= max(t_col, g_col): continue
            
            task_val = row[t_col]
            gestor_val = row[g_col]
            
            if task_val is not None and str(task_val).strip() != "":
                t_str = str(task_val).strip()
                t_str_lower = t_str.lower()
                
                if t_str_lower.startswith("set "):
                    current_set = t_str
                elif "cronograma" not in t_str_lower and gestor_val != "Gestor":
                    if gestor_val is not None and names_match(str(gestor_val).strip(), gestor_name):
                        assignments.append({
                            'set': current_set or 'Otros',
                            'task': t_str
                        })
    return resolved_shift, sheet_name, assignments

gestor = 'Juan Jose Diaz Alvarez'
for offset, day_name in [(0, "Viernes"), (1, "Sábado"), (2, "Domingo")]:
    dt = datetime.datetime(2026, 5, 22) + datetime.timedelta(days=offset)
    shift, sheet, assigns = test_date_assignments(gestor, dt)
    print(f"\nDate: {dt.strftime('%Y-%m-%d')} ({day_name})")
    print(f"  Shift: '{shift}'")
    print(f"  Sheet: '{sheet}'")
    print(f"  Assignments ({len(assigns)}):")
    for a in assigns:
        print(f"    - Set: '{a['set']}' | Task: '{a['task']}'")
